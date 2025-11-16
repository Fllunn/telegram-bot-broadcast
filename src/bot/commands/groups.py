from __future__ import annotations

import io
import logging
import math
import re
import secrets
from dataclasses import dataclass
from typing import Iterable, List, Mapping, Optional, Sequence
from urllib.parse import urlparse

import xlrd
from openpyxl import load_workbook
from telethon import Button, events, utils
from telethon.events import NewMessage
from telethon.tl.types import DocumentAttributeFilename

from src.bot.context import BotContext
from src.bot.keyboards import UPLOAD_GROUPS_LABEL, VIEW_GROUPS_LABEL, build_main_menu_keyboard
from src.models.session import TelethonSession
from src.services.broadcast_shared import (
    DialogsFetchError,
    collect_unique_target_peer_keys,
    deduplicate_broadcast_groups,
)
from src.services.groups_state import (
    GroupUploadScope,
    GroupUploadStateManager,
    GroupUploadStep,
    GroupViewSession,
    GroupViewScope,
    GroupViewStateManager,
    GroupViewStep,
    UploadAccountSnapshot,
)

logger = logging.getLogger(__name__)

CANCEL_LABEL = "Отмена"
UPLOAD_GROUPS_PATTERN = rf"^(?:/upload_groups(?:@\w+)?|{re.escape(UPLOAD_GROUPS_LABEL)})$"
VIEW_GROUPS_PATTERN = rf"^(?:/view_groups(?:@\w+)?|{re.escape(VIEW_GROUPS_LABEL)})$"
UPLOAD_SCOPE_PREFIX = "groups_scope"
UPLOAD_SCOPE_SINGLE = "single"
UPLOAD_SCOPE_ALL = "all"
SELECT_PREFIX = "groups_select"
CONFIRM_PREFIX = "groups_confirm"
CANCEL_PREFIX = "groups_cancel"
VIEW_SCOPE_PREFIX = "view_groups_scope"
VIEW_SCOPE_SINGLE = "single"
VIEW_SCOPE_ALL = "all"
VIEW_SELECT_PREFIX = "view_groups_select"
VIEW_PAGE_PREFIX = "view_groups_page"
VIEW_CANCEL_PREFIX = "view_groups_cancel"

ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
PAGE_SIZE = 10
PAYLOAD_VERSION = "v2"
DEDUP_NOTICE = (
    "Группы будут обработаны один раз для каждой уникальной группы на аккаунте. "
    "Количество строк в файле может отличаться от количества фактических рассылок."
)


@dataclass(frozen=True)
class ParsedGroup:
    name: Optional[str]
    username: Optional[str]
    link: Optional[str]


_USERNAME_TOKEN = re.compile(r"^(?:[A-Za-z0-9_]{3,64})$")


def _extract_username_candidate(value: str) -> Optional[str]:
    if not value:
        return None
    candidate = value.strip()
    if not candidate:
        return None
    candidate = candidate.lstrip("@")
    if not candidate:
        return None
    if _USERNAME_TOKEN.fullmatch(candidate):
        return candidate
    return None


def _normalize_link_value(value: str) -> Optional[str]:
    if not value:
        return None
    candidate = value.strip()
    if not candidate:
        return None
    identifier = _extract_identifier_from_link(candidate)
    if not identifier:
        return None
    return f"https://t.me/{identifier}"


def _prepare_group_fields(name: str, username: str, link: str) -> tuple[str, str, str]:
    fields = {
        "name": name.strip(),
        "username": username.strip(),
        "link": link.strip(),
    }
    for key in ("name", "username", "link"):
        value = fields.get(key) or ""
        if not value:
            continue
        link_candidate = _normalize_link_value(value)
        if link_candidate and not fields["link"]:
            fields["link"] = link_candidate
            if key != "link":
                fields[key] = ""
            continue
        username_candidate = _extract_username_candidate(value)
        if username_candidate and not fields["username"]:
            fields["username"] = username_candidate
            if key != "username":
                fields[key] = ""
    return fields["name"], fields["username"], fields["link"]


def _expect_step(context: BotContext, step: GroupUploadStep):
    def predicate(event: NewMessage.Event) -> bool:
        if not event.is_private or getattr(event.message, "out", False):
            return False
        state = context.groups_manager.get(event.sender_id)
        if state is None or state.step != step:
            return False
        if state.last_message_id is not None and state.last_message_id == event.id:
            return False
        return True

    return predicate


def _render_session_label(session: TelethonSession) -> str:
    display = session.display_name()
    phone = session.phone
    return f"{display} ({phone})" if phone else display


def _build_upload_snapshot(session: TelethonSession) -> UploadAccountSnapshot:
    return UploadAccountSnapshot(
        session_id=session.session_id,
        owner_id=session.owner_id,
        label=_render_session_label(session),
        cached_session=session,
    )


def _prepare_upload_snapshots(sessions: Iterable[TelethonSession]) -> dict[str, UploadAccountSnapshot]:
    snapshot_map: dict[str, UploadAccountSnapshot] = {}
    for session in sessions:
        snapshot = _build_upload_snapshot(session)
        snapshot_map[snapshot.session_id] = snapshot
    return snapshot_map


def _touch_snapshot(snapshot: UploadAccountSnapshot, session: TelethonSession) -> UploadAccountSnapshot:
    snapshot.cached_session = session
    snapshot.owner_id = session.owner_id
    snapshot.label = _render_session_label(session)
    return snapshot


def _join_session_identifier(parts: Sequence[str]) -> str:
    return ":".join(part for part in parts if part).strip()


def _build_upload_scope_buttons(flow_id: str) -> list[list[Button]]:
    return [
        [
            Button.inline(
                "Один аккаунт",
                f"{UPLOAD_SCOPE_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_SINGLE}".encode("utf-8"),
            ),
            Button.inline(
                "Все аккаунты",
                f"{UPLOAD_SCOPE_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_ALL}".encode("utf-8"),
            ),
        ],
        [
            Button.inline(
                "❌ Отмена",
                f"{CANCEL_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:scope".encode("utf-8"),
            )
        ],
    ]


def _build_upload_account_buttons(flow_id: str, sessions: Iterable[UploadAccountSnapshot]) -> list[list[Button]]:
    rows: list[list[Button]] = []
    for session in sessions:
        rows.append(
            [
                Button.inline(
                    session.label,
                    f"{SELECT_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{session.session_id}".encode("utf-8"),
                )
            ]
        )
    rows.append(
        [
            Button.inline(
                "❌ Отмена",
                f"{CANCEL_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:select".encode("utf-8"),
            )
        ]
    )
    return rows


def _build_upload_confirmation_buttons(
    flow_id: str,
    scope: GroupUploadScope,
    session_token: Optional[str] = None,
) -> list[list[Button]]:
    if scope == GroupUploadScope.SINGLE and session_token:
        yes_payload = (
            f"{CONFIRM_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_SINGLE}:yes:{session_token}".encode("utf-8")
        )
        no_payload = (
            f"{CONFIRM_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_SINGLE}:no:{session_token}".encode("utf-8")
        )
    else:
        yes_payload = f"{CONFIRM_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_ALL}:yes".encode("utf-8")
        no_payload = f"{CONFIRM_PREFIX}:{PAYLOAD_VERSION}:{flow_id}:{UPLOAD_SCOPE_ALL}:no".encode("utf-8")
    return [[Button.inline("✅ Да", yes_payload), Button.inline("❌ Нет", no_payload)]]


def _parse_callback_payload(data: bytes, prefix: str) -> Optional[tuple[Optional[str], list[str]]]:
    try:
        decoded = data.decode("utf-8")
    except UnicodeDecodeError:
        return None

    parts = decoded.split(":")
    if not parts or parts[0] != prefix:
        return None

    if len(parts) >= 3 and parts[1] == PAYLOAD_VERSION:
        flow_id = parts[2]
        remainder = parts[3:]
        return flow_id, remainder

    return None, parts[1:]


def _build_file_prompt_buttons() -> list[list[Button]]:
    return [[Button.text(CANCEL_LABEL, resize=True)]]


def _build_view_scope_buttons() -> list[list[Button]]:
    return [
        [
            Button.inline("Один аккаунт", f"{VIEW_SCOPE_PREFIX}:{VIEW_SCOPE_SINGLE}".encode("utf-8")),
            Button.inline("Все аккаунты", f"{VIEW_SCOPE_PREFIX}:{VIEW_SCOPE_ALL}".encode("utf-8")),
        ],
        [Button.inline("❌ Отмена", f"{VIEW_CANCEL_PREFIX}:scope".encode("utf-8"))],
    ]


def _build_view_account_buttons(sessions: Iterable[TelethonSession]) -> list[list[Button]]:
    rows: list[list[Button]] = []
    for session in sessions:
        rows.append(
            [
                Button.inline(
                    _render_session_label(session),
                    f"{VIEW_SELECT_PREFIX}:{session.session_id}".encode("utf-8"),
                )
            ]
        )
    rows.append([Button.inline("❌ Отмена", f"{VIEW_CANCEL_PREFIX}:select".encode("utf-8"))])
    return rows


def _normalize_cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def _parse_xlsx(content: bytes) -> List[ParsedGroup]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows: List[ParsedGroup] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        name = _normalize_cell_value(row[0]) if len(row) > 0 else ""
        username = _normalize_cell_value(row[1]) if len(row) > 1 else ""
        link = _normalize_cell_value(row[2]) if len(row) > 2 else ""
        name, username, link = _prepare_group_fields(name, username, link)
        if idx == 0 and _is_header_row(name, username, link):
            continue
        if not any((name, username, link)):
            continue
        rows.append(ParsedGroup(name=name or None, username=username or None, link=link or None))
    return rows


def _parse_xls(content: bytes) -> List[ParsedGroup]:
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    rows: List[ParsedGroup] = []
    for idx in range(sheet.nrows):
        row = sheet.row_values(idx)
        name = _normalize_cell_value(row[0]) if len(row) > 0 else ""
        username = _normalize_cell_value(row[1]) if len(row) > 1 else ""
        link = _normalize_cell_value(row[2]) if len(row) > 2 else ""
        name, username, link = _prepare_group_fields(name, username, link)
        if idx == 0 and _is_header_row(name, username, link):
            continue
        if not any((name, username, link)):
            continue
        rows.append(ParsedGroup(name=name or None, username=username or None, link=link or None))
    return rows


def _is_header_row(name: str, username: str, link: str) -> bool:
    header_tokens = {token.lower() for token in (name, username, link) if token}
    return bool(
        {"название", "название группы", "name"} & header_tokens
        or {"username", "юзернейм"} & header_tokens
        or {"ссылка", "link"} & header_tokens
    )



async def _resolve_chat_id(client, username: Optional[str], link: Optional[str]) -> tuple[Optional[int], Optional[bool]]:
    candidate = _sanitize_username(username) or _extract_identifier_from_link(link)
    if not candidate:
        return None, None

    try:
        entity = await client.get_input_entity(candidate)
    except Exception:
        return None, None

    try:
        peer_id = utils.get_peer_id(entity)
    except Exception:
        return None, None

    return peer_id, True


def _sanitize_username(username: Optional[str]) -> Optional[str]:
    if not username:
        return None
    username = username.strip()
    if not username:
        return None
    username = username.lstrip("@")
    if not username:
        return None
    return username


def _extract_identifier_from_link(link: Optional[str]) -> Optional[str]:
    if not link:
        return None
    link = link.strip()
    if not link:
        return None
    if link.startswith("http://") or link.startswith("https://"):
        parsed = urlparse(link)
        if parsed.netloc and parsed.netloc.lower().endswith("t.me"):
            path = parsed.path.lstrip("/")
            if path:
                return path.split("/", 1)[0]
        return None
    if link.startswith("t.me/"):
        return link.split("/", 1)[-1]
    return None


def _extract_filename(document) -> str:
    for attribute in document.attributes or []:
        if isinstance(attribute, DocumentAttributeFilename):
            return attribute.file_name or ""
    return ""


async def _parse_groups_file(file_bytes: bytes, extension: str) -> List[ParsedGroup]:
    if extension == ".xlsx":
        return _parse_xlsx(file_bytes)
    if extension == ".xls":
        return _parse_xls(file_bytes)
    raise ValueError("Unsupported file extension")


def _serialize_group(group: ParsedGroup, chat_id: Optional[int], is_member: Optional[bool]) -> Mapping[str, object]:
    return {
        "name": group.name,
        "username": group.username,
        "link": group.link,
        "chat_id": chat_id,
        "is_member": is_member,
    }


def _extract_groups(metadata: Optional[Mapping[str, object]]) -> List[Mapping[str, object]]:
    if not metadata:
        return []
    groups = metadata.get("broadcast_groups") if isinstance(metadata, Mapping) else None
    if isinstance(groups, list):
        normalized: List[Mapping[str, object]] = []
        for entry in groups:
            if isinstance(entry, Mapping):
                normalized.append(dict(entry))
        return normalized
    return []


def _purge_tokens_for_session(state: GroupViewSession, session_id: str) -> None:
    for token, stored_session in list(state.pagination_tokens.items()):
        if stored_session == session_id:
            state.pagination_tokens.pop(token, None)


def _register_pagination_token(state: GroupViewSession, session_id: str) -> str:
    token = secrets.token_hex(4)
    state.pagination_tokens[token] = session_id
    return token


def _build_view_pagination_buttons(state: GroupViewSession, session_id: str, page: int, total_pages: int) -> list[list[Button]]:
    _purge_tokens_for_session(state, session_id)
    buttons: list[list[Button]] = []
    nav_row: list[Button] = []
    if page > 0:
        prev_token = _register_pagination_token(state, session_id)
        nav_row.append(
            Button.inline(
                "‹ Предыдущие 10",
                f"{VIEW_PAGE_PREFIX}:{prev_token}:{page - 1}".encode("utf-8"),
            )
        )
    if page < total_pages - 1:
        next_token = _register_pagination_token(state, session_id)
        nav_row.append(
            Button.inline(
                "Следующие 10 ›",
                f"{VIEW_PAGE_PREFIX}:{next_token}:{page + 1}".encode("utf-8"),
            )
        )
    if nav_row:
        buttons.append(nav_row)
    buttons.append([Button.inline("❌ Закончить просмотр", f"{VIEW_CANCEL_PREFIX}:view".encode("utf-8"))])
    return buttons


def _format_group_entry(index: int, group: Mapping[str, object]) -> str:
    parts: list[str] = []

    name = group.get("name")
    if name is not None:
        name_value = str(name).strip()
        if name_value:
            parts.append(f"Название — \"{name_value}\"")

    username = group.get("username")
    if username is not None:
        username_value = str(username).strip()
        if username_value:
            parts.append(f"Username — @{username_value.lstrip('@')}")

    link = group.get("link")
    if link is not None:
        link_value = str(link).strip()
        if link_value:
            parts.append(f"Ссылка — {link_value}")

    if not parts:
        parts.append("Запись без данных")

    return f"{index}. " + ", ".join(parts)


def _format_groups_page(session: TelethonSession, groups: Sequence[Mapping[str, object]], page: int) -> str:
    label = _render_session_label(session)
    total = len(groups)
    if total == 0:
        return f"Аккаунт {label}\n\nДля этого аккаунта ещё не загружен список групп."

    total_pages = max(1, math.ceil(total / PAGE_SIZE))
    page = max(0, min(page, total_pages - 1))
    start = page * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)

    lines = [
        f"Аккаунт {label}",
        "",
        f"Группы для рассылки ({start + 1}-{end} из {total}):",
    ]

    for idx, group in enumerate(groups[start:end], start=start + 1):
        lines.append(_format_group_entry(idx, group))

    lines.append("")
    lines.append(f"Страница {page + 1} из {total_pages}.")
    return "\n".join(lines)


async def _handle_cancel(event: NewMessage.Event, manager: GroupUploadStateManager, message: str) -> None:
    user_id = event.sender_id
    manager.neutralize(user_id)
    manager.clear(user_id)
    await event.respond(message, buttons=build_main_menu_keyboard())


def setup_group_commands(client, context: BotContext) -> None:
    """Register commands for uploading and viewing broadcast group lists."""

    upload_manager = context.groups_manager
    view_manager = context.group_view_manager

    async def _get_available_sessions(user_id: int) -> Optional[List[TelethonSession]]:
        try:
            sessions = await context.session_repository.list_sessions_for_owner(user_id)
        except Exception:
            logger.exception(
                "Не удалось получить список аккаунтов",
                extra={"user_id": user_id},
            )
            return None

        if not sessions:
            return []

        # Prefer активные аккаунты, но позволяем пользователю выбирать и временно неактивные.
        sessions.sort(key=lambda item: (not item.is_active, item.display_name().lower()))
        return sessions

    async def _load_session_and_groups(session_id: str, cache: Mapping[str, TelethonSession]):
        session: Optional[TelethonSession] = None
        try:
            session = await context.session_repository.get_by_session_id(session_id)
        except Exception:
            logger.exception(
                "Не удалось загрузить данные аккаунта для просмотра групп",
                extra={"session_id": session_id},
            )
        if session is None:
            session = cache.get(session_id)
        if session is None:
            return None
        groups = _extract_groups(session.metadata)
        return session, groups

    async def _refresh_upload_state_sessions(user_id: int) -> Optional[object]:
        sessions = await _get_available_sessions(user_id)
        if sessions is None:
            return None
        snapshot_map = _prepare_upload_snapshots(sessions)
        state = upload_manager.update(
            user_id,
            sessions=snapshot_map,
            allowed_session_ids=list(snapshot_map.keys()),
        )
        return state

    async def _ensure_upload_snapshot(
        user_id: int,
        state,
        session_id: str,
        *,
        ensure_cached: bool,
    ) -> tuple[object, Optional[UploadAccountSnapshot]]:
        sessions_map = getattr(state, "sessions", None)
        if not isinstance(sessions_map, dict):
            sessions_map = {}
            setattr(state, "sessions", sessions_map)
        snapshot = sessions_map.get(session_id)
        if snapshot is None:
            refreshed_state = await _refresh_upload_state_sessions(user_id)
            if refreshed_state is None:
                return state, None
            state = refreshed_state
            sessions_map = getattr(state, "sessions", None)
            if not isinstance(sessions_map, dict):
                sessions_map = {}
                setattr(state, "sessions", sessions_map)
            snapshot = sessions_map.get(session_id)
        if snapshot is None:
            try:
                session_obj = await context.session_repository.get_by_session_id(session_id)
            except Exception:
                logger.exception(
                    "Не удалось получить данные аккаунта при проверке выбора",
                    extra={"session_id": session_id, "user_id": user_id},
                )
                session_obj = None
            if session_obj is not None and session_obj.owner_id == user_id:
                snapshot = _build_upload_snapshot(session_obj)
                sessions_map[session_id] = snapshot
                allowed = getattr(state, "allowed_session_ids", None)
                if not isinstance(allowed, list):
                    allowed = []
                    setattr(state, "allowed_session_ids", allowed)
                if session_id not in allowed:
                    allowed.append(session_id)
        if snapshot is None or snapshot.owner_id != user_id:
            return state, None
        if ensure_cached:
            cached = snapshot.cached_session
            if cached is None or cached.owner_id != user_id:
                try:
                    session_obj = await context.session_repository.get_by_session_id(session_id)
                except Exception:
                    logger.exception(
                        "Не удалось обновить данные аккаунта для загрузки групп",
                        extra={"session_id": session_id, "user_id": user_id},
                    )
                    return state, None
                if session_obj is None or session_obj.owner_id != user_id:
                    return state, None
                _touch_snapshot(snapshot, session_obj)
        return state, snapshot

    @client.on(events.NewMessage(pattern=UPLOAD_GROUPS_PATTERN))
    async def handle_upload_groups(event: NewMessage.Event) -> None:
        if not event.is_private:
            return

        user_id = event.sender_id
        existing_upload = upload_manager.get(user_id)
        if existing_upload and existing_upload.step != GroupUploadStep.IDLE:
            logger.info(
                "Сбрасываем незавершённую загрузку групп",
                extra={"user_id": user_id, "step": existing_upload.step.value},
            )
            upload_manager.clear(user_id)
        if upload_manager.has_active_flow(user_id):
            await event.respond(
                "Вы уже загружаете список групп. Завершите текущий процесс или отправьте «Отмена».",
                buttons=_build_file_prompt_buttons(),
            )
            return

        sessions = await _get_available_sessions(user_id)
        if sessions is None:
            await event.respond(
                "Не удалось получить список аккаунтов. Попробуйте позже.",
                buttons=build_main_menu_keyboard(),
            )
            return
        if not sessions:
            await event.respond(
                "У вас нет подключённых аккаунтов. Подключите аккаунт, чтобы загрузить список групп.",
                buttons=build_main_menu_keyboard(),
            )
            return

        snapshot_map = _prepare_upload_snapshots(sessions)
        state = upload_manager.begin(
            user_id,
            step=GroupUploadStep.CHOOSING_SCOPE,
            scope=GroupUploadScope.SINGLE,
            sessions=snapshot_map,
            allowed_session_ids=list(snapshot_map.keys()),
            last_message_id=event.id,
        )
        message = await event.respond(
            "Для каких аккаунтов загрузить список групп?",
            buttons=_build_upload_scope_buttons(state.flow_id),
        )
        upload_manager.update(user_id, last_message_id=message.id)

    @client.on(events.CallbackQuery(pattern=rf"^{UPLOAD_SCOPE_PREFIX}:".encode("utf-8")))
    async def handle_upload_scope_selection(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = upload_manager.get(user_id)
        if state is None or state.step != GroupUploadStep.CHOOSING_SCOPE:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        parsed = _parse_callback_payload(event.data, UPLOAD_SCOPE_PREFIX)
        if parsed is None:
            await event.answer("Некорректный запрос.", alert=True)
            return

        flow_id, parts = parsed
        if not parts:
            await event.answer("Некорректный выбор.", alert=True)
            return
        selection = parts[0]

        if flow_id != state.flow_id:
            await event.answer("Сценарий загрузки устарел. Запустите /upload_groups заново.", alert=True)
            return

        sessions = list(state.sessions.values()) if state.sessions else []

        if selection == UPLOAD_SCOPE_SINGLE:
            if not sessions:
                upload_manager.clear(user_id)
                await event.edit("Нет доступных аккаунтов для загрузки.", buttons=build_main_menu_keyboard())
                return
            upload_manager.reset_targets(user_id)
            upload_manager.update(user_id, scope=GroupUploadScope.SINGLE, step=GroupUploadStep.CHOOSING_ACCOUNT)
            message = await event.edit(
                "Выберите аккаунт, для которого нужно загрузить список групп.",
                buttons=_build_upload_account_buttons(state.flow_id, sessions),
            )
            upload_manager.update(user_id, last_message_id=message.id)
            return

        if selection == UPLOAD_SCOPE_ALL:
            if not sessions:
                upload_manager.clear(user_id)
                await event.edit("Нет доступных аккаунтов для загрузки.", buttons=build_main_menu_keyboard())
                return
            session_ids = [session.session_id for session in sessions]
            upload_manager.reset_targets(user_id)
            upload_manager.set_all_targets(user_id, session_ids)
            has_existing = any(
                snapshot.cached_session is not None and _extract_groups(snapshot.cached_session.metadata)
                for snapshot in sessions
            )
            if has_existing:
                upload_manager.update(user_id, step=GroupUploadStep.CONFIRMING_REPLACE)
                message = await event.edit(
                    "В некоторых аккаунтах уже есть список групп. Заменить его для всех аккаунтов?",
                    buttons=_build_upload_confirmation_buttons(state.flow_id, GroupUploadScope.ALL),
                )
            else:
                upload_manager.update(user_id, step=GroupUploadStep.WAITING_FILE)
                message = await event.edit(
                    "Отправьте Excel-файл (.xlsx или .xls) со списком групп. Первая строка может быть заголовком.",
                    buttons=_build_file_prompt_buttons(),
                )
            upload_manager.update(user_id, last_message_id=message.id)
            return

        await event.answer("Некорректный выбор.", alert=True)

    @client.on(events.CallbackQuery(pattern=rf"^{SELECT_PREFIX}:".encode("utf-8")))
    async def handle_upload_account_selection(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = upload_manager.get(user_id)
        if state is None or state.step != GroupUploadStep.CHOOSING_ACCOUNT:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        parsed = _parse_callback_payload(event.data, SELECT_PREFIX)
        if parsed is None:
            await event.answer("Некорректный запрос.", alert=True)
            return

        flow_id, parts = parsed
        if not parts:
            await event.answer("Не удалось определить аккаунт. Запустите загрузку заново.", alert=True)
            return
        session_id = _join_session_identifier(parts)
        if not session_id:
            await event.answer("Не удалось определить аккаунт. Запустите загрузку заново.", alert=True)
            return

        if flow_id != state.flow_id:
            await event.answer("Сценарий загрузки устарел. Запустите /upload_groups заново.", alert=True)
            return

        if not state.sessions:
            refreshed_state = await _refresh_upload_state_sessions(user_id)
            if refreshed_state is None or not getattr(refreshed_state, "sessions", None):
                upload_manager.clear(user_id)
                await event.answer("Список аккаунтов устарел. Начните загрузку заново.", alert=True)
                return
            state = refreshed_state

        state, snapshot = await _ensure_upload_snapshot(user_id, state, session_id, ensure_cached=True)
        if snapshot is None:
            upload_manager.reset_targets(user_id)
            await event.answer("Не удалось подтвердить выбранный аккаунт. Запустите загрузку заново.", alert=True)
            return

        upload_manager.reset_targets(user_id)
        updated_state = upload_manager.set_single_target(user_id, session_id)
        if updated_state is None:
            await event.answer("Выбор аккаунта устарел. Начните загрузку заново.", alert=True)
            return
        state = updated_state

        session_obj = snapshot.cached_session
        existing = _extract_groups(session_obj.metadata) if session_obj is not None else []
        if existing:
            token = upload_manager.register_confirmation_token(user_id, session_id)
            if not token:
                await event.answer("Не удалось подготовить подтверждение. Повторите попытку.", alert=True)
                return
            upload_manager.update(user_id, step=GroupUploadStep.CONFIRMING_REPLACE)
            message = await event.edit(
                "Для выбранного аккаунта уже есть список групп. Заменить его?",
                buttons=_build_upload_confirmation_buttons(state.flow_id, GroupUploadScope.SINGLE, token),
            )
        else:
            upload_manager.update(user_id, step=GroupUploadStep.WAITING_FILE)
            message = await event.edit(
                "Отправьте Excel-файл (.xlsx или .xls) со списком групп. Первая строка может быть заголовком.",
                buttons=_build_file_prompt_buttons(),
            )
        upload_manager.update(user_id, last_message_id=message.id)

    @client.on(events.CallbackQuery(pattern=rf"^{CONFIRM_PREFIX}:".encode("utf-8")))
    async def handle_upload_confirmation(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = upload_manager.get(user_id)
        if state is None or state.step != GroupUploadStep.CONFIRMING_REPLACE:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        parsed = _parse_callback_payload(event.data, CONFIRM_PREFIX)
        if parsed is None:
            await event.answer("Некорректный запрос.", alert=True)
            return

        flow_id, parts = parsed
        if len(parts) < 2:
            await event.answer("Некорректный запрос.", alert=True)
            return

        scope_marker, decision, *rest = parts
        scope = GroupUploadScope.ALL if scope_marker == UPLOAD_SCOPE_ALL else GroupUploadScope.SINGLE

        if flow_id != state.flow_id:
            await event.answer("Сценарий загрузки устарел. Запустите /upload_groups заново.", alert=True)
            return

        if decision == "no":
            upload_manager.neutralize(user_id)
            upload_manager.clear(user_id)
            await event.edit("Загрузка списка групп отменена.", buttons=build_main_menu_keyboard())
            return

        if decision != "yes":
            await event.answer("Некорректный выбор.", alert=True)
            return

        if scope == GroupUploadScope.SINGLE:
            if not rest:
                await event.answer("Некорректный запрос.", alert=True)
                return
            session_token = rest[0].strip()
            if not session_token:
                await event.answer("Некорректный запрос.", alert=True)
                return
            session_id = upload_manager.consume_confirmation_token(user_id, session_token)
            selected_id = getattr(state, "selected_session_id", None)
            target_ids = list(state.target_session_ids or [])
            if not session_id or not selected_id or not target_ids:
                upload_manager.reset_targets(user_id)
                await event.answer("Аккаунт не выбран. Повторите выбор.", alert=True)
                return
            if session_id != selected_id or session_id not in target_ids:
                upload_manager.reset_targets(user_id)
                await event.answer("Выбор аккаунта устарел. Выберите аккаунт заново.", alert=True)
                return
            if session_id not in state.sessions:
                upload_manager.clear(user_id)
                await event.answer("Список аккаунтов устарел. Запустите загрузку заново.", alert=True)
                return

        upload_manager.update(user_id, step=GroupUploadStep.WAITING_FILE)
        message = await event.edit(
            "Отправьте Excel-файл (.xlsx или .xls) со списком групп. Первая строка может быть заголовком.",
            buttons=_build_file_prompt_buttons(),
        )
        upload_manager.update(user_id, last_message_id=message.id)

    @client.on(events.CallbackQuery(pattern=rf"^{CANCEL_PREFIX}:".encode("utf-8")))
    async def handle_upload_inline_cancel(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = upload_manager.get(user_id)
        if state is None:
            await event.answer("Нечего отменять.", alert=True)
            return

        parsed = _parse_callback_payload(event.data, CANCEL_PREFIX)
        if parsed is None:
            await event.answer("Некорректный запрос.", alert=True)
            return

        flow_id, _ = parsed
        if flow_id != state.flow_id:
            await event.answer("Сценарий загрузки устарел. Запустите /upload_groups заново.", alert=True)
            return

        if not upload_manager.has_active_flow(user_id):
            await event.answer("Нечего отменять.", alert=True)
            return
        upload_manager.neutralize(user_id)
        upload_manager.clear(user_id)
        await event.edit("Загрузка списка групп отменена.", buttons=build_main_menu_keyboard())

    @client.on(events.NewMessage(incoming=True, func=_expect_step(context, GroupUploadStep.WAITING_FILE)))
    async def handle_upload_file(event: NewMessage.Event) -> None:
        user_id = event.sender_id
        message_text = (event.raw_text or "").strip()
        if message_text.lower() == CANCEL_LABEL.lower():
            await _handle_cancel(event, upload_manager, "Загрузка списка групп отменена.")
            return

        document = event.document
        if document is None:
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Пожалуйста, отправьте Excel-файл формата .xlsx или .xls, либо напишите «Отмена».",
                buttons=_build_file_prompt_buttons(),
            )
            return

        filename = _extract_filename(document)
        extension = _detect_extension(document.mime_type, filename)
        if extension not in ALLOWED_EXTENSIONS:
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Формат файла не поддерживается. Отправьте Excel-файл (.xlsx или .xls) или напишите «Отмена».",
                buttons=_build_file_prompt_buttons(),
            )
            return

        try:
            file_bytes = await event.download_media(bytes)
        except Exception:
            logger.exception("Не удалось скачать файл со списком групп", extra={"user_id": user_id})
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Не удалось скачать файл. Попробуйте ещё раз или отправьте «Отмена».",
                buttons=_build_file_prompt_buttons(),
            )
            return

        try:
            parsed_groups = await _parse_groups_file(file_bytes, extension)
        except Exception:
            logger.exception("Ошибка при чтении Excel с группами", extra={"user_id": user_id})
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Не удалось прочитать файл. Убедитесь, что это корректный Excel (.xlsx или .xls).",
                buttons=_build_file_prompt_buttons(),
            )
            return

        if not parsed_groups:
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Файл не содержит строк со списком групп. Заполните хотя бы одно поле в каждой строке.",
                buttons=_build_file_prompt_buttons(),
            )
            return

        state = upload_manager.get(user_id)
        if state is None:
            logger.warning("Состояние загрузки групп потеряно", extra={"user_id": user_id})
            await _handle_cancel(event, upload_manager, "Не удалось определить целевые аккаунты. Попробуйте снова.")
            return

        target_ids = list(state.target_session_ids or [])
        resolved_snapshots: list[UploadAccountSnapshot] = []

        if state.scope == GroupUploadScope.SINGLE:
            selected_id = getattr(state, "selected_session_id", None)
            if not selected_id:
                logger.warning("Нет выбранного аккаунта при загрузке одиночного списка групп", extra={"user_id": user_id})
                await _handle_cancel(event, upload_manager, "Аккаунт не выбран. Запустите загрузку заново.")
                return
            state, snapshot = await _ensure_upload_snapshot(user_id, state, selected_id, ensure_cached=True)
            if snapshot is None:
                logger.warning(
                    "Не удалось подтвердить выбранный аккаунт при загрузке файла",
                    extra={"user_id": user_id, "session_id": selected_id},
                )
                await _handle_cancel(event, upload_manager, "Выбор аккаунта устарел. Запустите загрузку заново.")
                return
            target_ids = [snapshot.session_id]
            resolved_snapshots.append(snapshot)
        else:
            if not target_ids:
                logger.warning("Нет целевых аккаунтов для сохранения групп", extra={"user_id": user_id})
                await _handle_cancel(event, upload_manager, "Не удалось определить целевые аккаунты. Попробуйте снова.")
                return
            sanitized_ids: list[str] = []
            for session_id in target_ids:
                state, snapshot = await _ensure_upload_snapshot(user_id, state, session_id, ensure_cached=True)
                if snapshot is None:
                    logger.warning(
                        "Целевой аккаунт недоступен при загрузке файла",
                        extra={"user_id": user_id, "session_id": session_id},
                    )
                    await _handle_cancel(event, upload_manager, "Список аккаунтов устарел. Запустите загрузку заново.")
                    return
                sanitized_ids.append(snapshot.session_id)
                resolved_snapshots.append(snapshot)
            target_ids = sanitized_ids

        enriched_groups: list[dict[str, object]] = []
        for group in parsed_groups:
            chat_id, is_member = await _resolve_chat_id(event.client, group.username, group.link)
            enriched_groups.append(_serialize_group(group, chat_id, is_member))

        unique_groups = deduplicate_broadcast_groups(enriched_groups)
        groups_stats = {
            "file_rows": len(enriched_groups),
            "unique_groups": len(unique_groups),
        }
        account_stats: dict[str, dict[str, object]] = {}
        for snapshot in resolved_snapshots:
            stats_for_account = dict(groups_stats)
            session_obj = snapshot.cached_session
            if session_obj is None:
                try:
                    session_obj = await context.session_repository.get_by_session_id(snapshot.session_id)
                except Exception:
                    logger.exception(
                        "Не удалось загрузить данные аккаунта при расчёте групп",
                        extra={"user_id": user_id, "session_id": snapshot.session_id},
                    )
                    session_obj = None
                else:
                    snapshot.cached_session = session_obj
            actual_targets = len(unique_groups)
            if session_obj is not None and unique_groups:
                session_client = None
                try:
                    session_client = await context.session_manager.build_client_from_session(session_obj)
                    peer_keys = await collect_unique_target_peer_keys(
                        session_client,
                        unique_groups,
                        user_id=user_id,
                        account_label=snapshot.label,
                        account_session_id=session_obj.session_id,
                    )
                    actual_targets = len(peer_keys)
                except DialogsFetchError as exc:
                    logger.warning(
                        "Не удалось проверить список чатов при загрузке групп",
                        extra={"user_id": user_id, "session_id": snapshot.session_id, "reason": exc.error_type},
                    )
                except Exception:
                    logger.exception(
                        "Ошибка при расчёте фактических групп при загрузке",
                        extra={"user_id": user_id, "session_id": snapshot.session_id},
                    )
                finally:
                    if session_client is not None:
                        try:
                            await context.session_manager.close_client(session_client)
                        except Exception:
                            logger.exception(
                                "Не удалось закрыть клиент после расчёта фактических групп",
                                extra={"session_id": snapshot.session_id},
                            )
            stats_for_account["actual_targets"] = actual_targets
            account_stats[snapshot.session_id] = stats_for_account
        for session_id in target_ids:
            if session_id not in account_stats:
                stats_fallback = dict(groups_stats)
                stats_fallback["actual_targets"] = len(unique_groups)
                account_stats[session_id] = stats_fallback

        snapshot_lookup = {snapshot.session_id: snapshot for snapshot in resolved_snapshots}
        operation_scope = state.scope
        try:
            if operation_scope == GroupUploadScope.ALL:
                updated = 0
                for session_id in target_ids:
                    snapshot = snapshot_lookup.get(session_id)
                    stats_for_account = account_stats.get(session_id, dict(groups_stats))
                    success = await context.session_repository.set_broadcast_groups(
                        session_id,
                        enriched_groups,
                        owner_id=user_id,
                        unique_groups=unique_groups,
                        stats=stats_for_account,
                    )
                    if not success:
                        label = snapshot.label if snapshot else session_id
                        raise RuntimeError(f"Не удалось обновить аккаунт {label}")
                    updated += 1
                if updated != len(target_ids):
                    raise RuntimeError("Не все аккаунты подтвердили обновление списка групп")
                upload_manager.reset_targets(user_id)
            else:
                session_id = target_ids[0]
                stats_for_account = account_stats.get(session_id, dict(groups_stats))
                success = await context.session_repository.set_broadcast_groups(
                    session_id,
                    enriched_groups,
                    owner_id=user_id,
                    unique_groups=unique_groups,
                    stats=stats_for_account,
                )
                if not success:
                    raise RuntimeError("Не удалось обновить выбранный аккаунт")
        except Exception:
            logger.exception(
                "Ошибка при сохранении списка групп",
                extra={"user_id": user_id, "scope": state.scope.value, "targets": target_ids},
            )
            upload_manager.update(user_id, last_message_id=event.id)
            await event.respond(
                "Не удалось сохранить список групп. Попробуйте позже или отправьте «Отмена».",
                buttons=_build_file_prompt_buttons(),
            )
            return

        if operation_scope == GroupUploadScope.ALL:
            success_text = "Список групп для рассылки успешно загружен для всех подключённых аккаунтов."
        else:
            snapshot = resolved_snapshots[0] if resolved_snapshots else None
            label = snapshot.label if snapshot else "выбранного аккаунта"
            success_text = f"Список групп для аккаунта {label} успешно обновлён."

        success_text = f"{success_text}\n\n{DEDUP_NOTICE}"

        total_actual_targets = 0
        for stats_payload in account_stats.values():
            value = stats_payload.get("actual_targets")
            try:
                total_actual_targets += int(value)
            except (TypeError, ValueError):
                total_actual_targets += len(unique_groups)
        logger.info(
            "Пользователь %s загрузил список групп (scope=%s, rows=%s, unique=%s, actual=%s)",
            user_id,
            state.scope.value,
            len(enriched_groups),
            len(unique_groups),
            total_actual_targets,
        )

        upload_manager.neutralize(user_id)
        upload_manager.clear(user_id)
        await event.respond(success_text, buttons=build_main_menu_keyboard())

    @client.on(events.NewMessage(pattern=VIEW_GROUPS_PATTERN))
    async def handle_view_groups(event: NewMessage.Event) -> None:
        if not event.is_private:
            return

        user_id = event.sender_id
        existing_view = view_manager.get(user_id)
        if existing_view and existing_view.step != GroupViewStep.IDLE:
            logger.info(
                "Сбрасываем незавершённый просмотр групп",
                extra={"user_id": user_id, "step": existing_view.step.value},
            )
            view_manager.clear(user_id)
        if view_manager.has_active_flow(user_id):
            await event.respond(
                "Вы уже просматриваете списки групп. Завершите текущий просмотр или используйте кнопку «❌ Закончить просмотр».",
                buttons=build_main_menu_keyboard(),
            )
            return

        sessions = await _get_available_sessions(user_id)
        if sessions is None:
            await event.respond(
                "Не удалось получить список аккаунтов. Попробуйте позже.",
                buttons=build_main_menu_keyboard(),
            )
            return
        if not sessions:
            await event.respond(
                "У вас нет подключённых аккаунтов. Подключите аккаунт, чтобы просматривать списки групп.",
                buttons=build_main_menu_keyboard(),
            )
            return

        view_manager.begin(
            user_id,
            step=GroupViewStep.CHOOSING_SCOPE,
            session_ids=[session.session_id for session in sessions],
            sessions={session.session_id: session for session in sessions},
            last_message_id=event.id,
        )
        message = await event.respond(
            "Для каких аккаунтов показать сохранённые группы?",
            buttons=_build_view_scope_buttons(),
        )
        view_manager.update(user_id, last_message_id=message.id)

    @client.on(events.CallbackQuery(pattern=rf"^{VIEW_SCOPE_PREFIX}:".encode("utf-8")))
    async def handle_view_scope_selection(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = view_manager.get(user_id)
        if state is None or state.step != GroupViewStep.CHOOSING_SCOPE:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        selection = event.data.decode("utf-8").split(":", maxsplit=1)[-1]
        sessions = list(state.sessions.values())

        if selection == VIEW_SCOPE_SINGLE:
            if not sessions:
                view_manager.clear(user_id)
                await event.edit("Нет доступных аккаунтов для просмотра.", buttons=build_main_menu_keyboard())
                return
            view_manager.update(user_id, scope=GroupViewScope.SINGLE, step=GroupViewStep.CHOOSING_ACCOUNT)
            message = await event.edit(
                "Выберите аккаунт, для которого показать сохранённые группы.",
                buttons=_build_view_account_buttons(sessions),
            )
            view_manager.update(user_id, last_message_id=message.id)
            return

        if selection == VIEW_SCOPE_ALL:
            if not sessions:
                view_manager.clear(user_id)
                await event.edit("Нет доступных аккаунтов для просмотра.", buttons=build_main_menu_keyboard())
                return
            view_manager.update(user_id, scope=GroupViewScope.ALL, step=GroupViewStep.VIEWING)
            await event.edit(
                "Показываю списки групп для всех подключённых аккаунтов. Используйте кнопки под сообщениями для навигации.",
            )
            for session in sessions:
                loaded = await _load_session_and_groups(session.session_id, state.sessions)
                if loaded is None:
                    await event.client.send_message(
                        user_id,
                        f"Не удалось получить данные для аккаунта {_render_session_label(session)}. Попробуйте позже.",
                    )
                    continue
                session_obj, groups = loaded
                state.sessions[session_obj.session_id] = session_obj
                total_pages = max(1, math.ceil(len(groups) / PAGE_SIZE))
                text = _format_groups_page(session_obj, groups, 0)
                buttons = _build_view_pagination_buttons(state, session_obj.session_id, 0, total_pages)
                await event.client.send_message(user_id, text, buttons=buttons)
            return

        await event.answer("Некорректный выбор.", alert=True)

    @client.on(events.CallbackQuery(pattern=rf"^{VIEW_SELECT_PREFIX}:".encode("utf-8")))
    async def handle_view_account_selection(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        state = view_manager.get(user_id)
        if state is None or state.step != GroupViewStep.CHOOSING_ACCOUNT:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        session_id = event.data.decode("utf-8").split(":", maxsplit=1)[-1]
        loaded = await _load_session_and_groups(session_id, state.sessions)
        if loaded is None:
            view_manager.clear(user_id)
            await event.edit(
                "Не удалось получить данные выбранного аккаунта. Попробуйте позже.",
                buttons=build_main_menu_keyboard(),
            )
            return

        session_obj, groups = loaded
        state.sessions[session_obj.session_id] = session_obj
        view_manager.update(
            user_id,
            scope=GroupViewScope.SINGLE,
            step=GroupViewStep.VIEWING,
            session_ids=[session_obj.session_id],
        )
        await event.edit(
            f"Аккаунт {_render_session_label(session_obj)} выбран. Отображаю список групп.",
        )
        total_pages = max(1, math.ceil(len(groups) / PAGE_SIZE))
        text = _format_groups_page(session_obj, groups, 0)
        buttons = _build_view_pagination_buttons(state, session_obj.session_id, 0, total_pages)
        await event.client.send_message(user_id, text, buttons=buttons)

    @client.on(events.CallbackQuery(pattern=rf"^{VIEW_PAGE_PREFIX}:".encode("utf-8")))
    async def handle_view_pagination(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        try:
            text = event.data.decode("utf-8")
        except UnicodeDecodeError:
            await event.answer("Некорректный запрос.", alert=True)
            return

        parts = text.split(":", maxsplit=2)
        if len(parts) != 3 or parts[0] != VIEW_PAGE_PREFIX:
            await event.answer("Некорректный запрос.", alert=True)
            return

        _, token, page_str = parts
        state = view_manager.get(user_id)
        if state is None:
            await event.answer("Сеанс просмотра устарел. Начните заново командой /view_groups.", alert=True)
            return

        session_id = state.pagination_tokens.pop(token, None)
        if session_id is None:
            logger.warning(
                "Получен неизвестный токен пагинации",
                extra={"sender_id": user_id, "token": token},
            )
            await event.answer("Сеанс просмотра устарел. Начните заново командой /view_groups.", alert=True)
            return

        if state.step != GroupViewStep.VIEWING:
            await event.answer("Эта операция больше неактуальна.", alert=True)
            return

        loaded = await _load_session_and_groups(session_id, state.sessions)
        if loaded is None:
            await event.answer("Не удалось получить данные аккаунта.", alert=True)
            return

        session_obj, groups = loaded
        if session_obj.owner_id != user_id:
            logger.warning(
                "Попытка просмотра списка групп чужого аккаунта",
                extra={"sender_id": user_id, "owner_id": session_obj.owner_id, "session_id": session_obj.session_id},
            )
            await event.answer("Нет доступа к этому аккаунту.", alert=True)
            return

        state.sessions[session_obj.session_id] = session_obj
        if session_obj.session_id not in state.session_ids:
            state.session_ids.append(session_obj.session_id)

        try:
            requested_page = int(page_str)
        except ValueError:
            await event.answer("Некорректный запрос.", alert=True)
            return

        total_pages = max(1, math.ceil(len(groups) / PAGE_SIZE))
        page = max(0, min(requested_page, total_pages - 1))
        text = _format_groups_page(session_obj, groups, page)
        buttons = _build_view_pagination_buttons(state, session_obj.session_id, page, total_pages)
        await event.edit(text, buttons=buttons)
        await event.answer()

    @client.on(events.CallbackQuery(pattern=rf"^{VIEW_CANCEL_PREFIX}:".encode("utf-8")))
    async def handle_view_cancel(event: events.CallbackQuery.Event) -> None:
        user_id = event.sender_id
        if not view_manager.has_active_flow(user_id):
            await event.answer("Нечего отменять.", alert=True)
            return
        view_manager.clear(user_id)
        await event.edit("Просмотр списков групп завершён.", buttons=build_main_menu_keyboard())


def _detect_extension(mime_type: Optional[str], filename: str) -> str:
    if filename:
        lowered = filename.lower()
        for ext in ALLOWED_EXTENSIONS:
            if lowered.endswith(ext):
                return ext
    if mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return ".xlsx"
    if mime_type == "application/vnd.ms-excel":
        return ".xls"
    return ""
